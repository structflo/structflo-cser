"""Streaming/serialisation check for the web UI — no models are loaded."""

import io
import json

import pytest
from PIL import Image

from structflo.cser.pipeline import BBox, CompoundPair, Detection
from webapp import server


def _pair(smiles, text):
    det = lambda cid: Detection(bbox=BBox(10, 10, 60, 60), conf=0.9, class_id=cid)  # noqa: E731
    return CompoundPair(
        structure=det(0),
        label=det(1),
        match_distance=1.0,
        smiles=smiles,
        label_text=text,
        match_confidence=0.87,
    )


class FakePipeline:
    def process(self, img):
        return [_pair("CCO", "1a"), _pair("not-a-smiles", "1b")]


@pytest.fixture
def client(monkeypatch):
    monkeypatch.setattr(server, "_pipeline", FakePipeline())
    return server.app.test_client()


def _png_bytes():
    buf = io.BytesIO()
    Image.new("RGB", (200, 200), "white").save(buf, "PNG")
    return buf.getvalue()


def test_extract_streams_pairs(client):
    res = client.post("/extract", data={"file": (io.BytesIO(_png_bytes()), "page.png")})
    assert res.status_code == 200

    lines = [
        json.loads(x) for x in res.get_data(as_text=True).splitlines() if x.strip()
    ]
    meta, page = lines[0], lines[1]
    assert meta["n_pages"] == 1
    assert page["w"] == 200 and page["image"].startswith("data:image/jpeg;base64,")

    ok, bad = page["pairs"]
    assert (ok["id"], ok["smiles"]) == ("1a", "CCO")
    assert ok["smiles_img"].startswith("data:image/png;base64,")  # rdkit re-render
    assert bad["smiles_img"] is None  # unparseable SMILES
    assert ok["structure_bbox"] == [10, 10, 60, 60]


def test_rejects_unsupported_file(client):
    res = client.post("/extract", data={"file": (io.BytesIO(b"x"), "notes.txt")})
    assert res.status_code == 400


def test_xlsx_export_keeps_ids_as_text(client):
    res = client.post(
        "/export.xlsx",
        json={
            "name": "paper.pdf",
            "rows": [
                {"page": 1, "id": "7178-39-6", "smiles": "CCO", "confidence": 0.9}
            ],
        },
    )
    assert res.status_code == 200
    assert "paper.xlsx" in res.headers["Content-Disposition"]

    from openpyxl import load_workbook

    ws = load_workbook(io.BytesIO(res.get_data())).active
    assert [c.value for c in ws[1]] == ["ID", "SMILES", "Page", "Match confidence"]
    # the whole point of xlsx over CSV: Excel would read "7178-39-6" as a date
    assert ws["A2"].value == "7178-39-6" and isinstance(ws["A2"].value, str)


def test_xlsx_export_rejects_bad_payload(client):
    assert client.post("/export.xlsx", json={"rows": "nope"}).status_code == 400
